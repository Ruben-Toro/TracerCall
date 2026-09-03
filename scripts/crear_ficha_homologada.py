from __future__ import annotations

import shutil
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo


SOURCE = Path(r"D:\Documentos\1 Archivos - Ruben Dario Toro Torres\Innovate\20260823 Demo Sophia\Demo Sophia Innovate.xlsx")
OUT_DIR = Path(r"D:\Documentos\ChatGPT\Generador de Traza\outputs\01a06799-e229-7010-9bdf-d9cbf8ab159d")
OUTPUT = OUT_DIR / "Ficha_maestra_homologada_SophIA.xlsx"


FIELDS = [
    ("Control", "0.1", "Nombre del agente", "name", "string", "Andres", "Campo API", "Sí", "Confirmado", "Nombre visible del agente."),
    ("Control", "0.2", "Canales habilitados", "enabled_channels", "array<string>", "voice", "Campo API", "Sí", "Confirmado", "Enviar como arreglo: [\"voice\"]."),
    ("Identidad", "1.1", "Mensaje inicial", "initialMessage", "string", "Hola soy Andres, un agente especializado de Innovate Solutions", "Campo API", "Sí", "Confirmado", "Debe coincidir con la apertura del prompt."),
    ("Identidad", "1.2", "Rol del agente", "systemPrompt", "texto: IDENTIDAD.rol", "Agente especializado", "Prompt", "Sí", "Confirmado", "No existe como columna independiente."),
    ("Identidad", "1.3", "Empresa representada", "systemPrompt", "texto: IDENTIDAD.empresa", "Innovate", "Prompt", "Sí", "Confirmado", "No existe como columna independiente."),
    ("Identidad", "1.4", "Usuario objetivo", "systemPrompt", "texto: IDENTIDAD.usuario_objetivo", "Clientes con agenda para realizar una demostración", "Prompt", "Sí", "Pendiente", "El PDF también dice Clientes de SophIA; escoger un único valor."),
    ("Identidad", "1.5", "Propósito del agente", "systemPrompt", "texto: OBJETIVOS.proposito", "Orientar y ayudar a Innovate con información relevante para crear una demostración acorde a las necesidades reales", "Prompt", "Sí", "Confirmado", "No existe como columna independiente."),
    ("Identidad", "1.6", "Transparencia sobre IA", "systemPrompt", "texto: TRANSPARENCIA_IA", "Aclarar que es una IA cuando el usuario lo pregunte", "Prompt", "Sí", "Confirmado", "Regla condicional."),
    ("Idioma", "1.7", "Idioma de conversación y STT", "language", "locale", "es-419", "Campo API", "Sí", "Confirmado", "Homologación de Español latinoamericano."),
    ("Identidad", "1.8", "Tratamiento al usuario", "systemPrompt", "texto: TRATAMIENTO", "Usted", "Prompt", "Sí", "Confirmado", "Consolida los dos ítems duplicados del PDF."),
    ("Voz", "1.9", "Voz del agente", "voice", "integer (FK)", "Pablo / voz masculina", "Campo API", "Sí", "Pendiente", "Resolver el nombre comercial al ID vigente."),
    ("Voz", "1.10", "Proveedor TTS", "ttsProvider", "integer (FK)", "iMash TTS", "Campo API", "Sí", "Pendiente", "Resolver al ID vigente del proveedor."),
    ("Voz", "1.11", "Modelo TTS", "ttsModel", "integer|string|null", "", "Campo API", "Sí", "Pendiente", "Completar si la voz seleccionada exige modelo."),
    ("Voz", "1.12", "Velocidad de voz", "voiceCustomSettings.speed", "number", "0.3", "Objeto API", "Sí", "Pendiente", "El PDF mezcla este valor con LLM; confirmar antes de publicar."),
    ("LLM", "1.13", "Temperatura del modelo", "llmTemperature", "number", "", "Campo API", "Sí", "Pendiente", "No reutilizar 0.3 hasta aclarar si era velocidad o temperatura."),
    ("Estilo", "1.14", "Personalidad", "systemPrompt", "texto: ESTILO.personalidad", "Profesional, empático y confiable", "Prompt", "Sí", "Confirmado", "Consolida el duplicado del PDF."),
    ("Estilo", "1.15", "Nivel de formalidad", "systemPrompt", "texto: ESTILO.formalidad", "Profesional confiable", "Prompt", "Sí", "Confirmado", "Consolida el duplicado del PDF."),
    ("Estilo", "1.16", "Tono", "systemPrompt", "texto: ESTILO.tono", "Emocionado por crear una excelente demostración acorde a las necesidades", "Prompt", "Sí", "Confirmado", "Regla de estilo."),
    ("Estilo", "1.17", "Uso de humor", "systemPrompt", "texto: ESTILO.humor", "No utilizar humor", "Prompt", "Sí", "Confirmado", "Regla de estilo."),
    ("Estilo", "1.18", "Uso de regionalismos", "systemPrompt", "texto: ESTILO.regionalismos", "No usar regionalismos", "Prompt", "Sí", "Confirmado", "Regla de estilo."),
    ("Lenguaje", "1.19", "Lenguaje permitido", "systemPrompt", "texto: LENGUAJE.permitido", "Usar lenguaje profesional", "Prompt", "Sí", "Confirmado", "Reemplaza Palabras permitidas."),
    ("Lenguaje", "1.20", "Lenguaje prohibido", "systemPrompt", "texto: LENGUAJE.prohibido", "Evitar imposible, no puedo, no se puede; evitar metáforas y exageraciones", "Prompt", "Sí", "Confirmado", "Reemplaza Palabras prohibidas."),
    ("Lenguaje", "1.21", "Pronunciación de números", "systemPrompt", "texto: PRONUNCIACION.numeros", "Leer en miles o redondear a cientos", "Prompt", "Sí", "Confirmado", "Puede complementarse con speechNormalization."),
    ("Lenguaje", "1.22", "Nivel de certeza", "systemPrompt", "texto: CERTEZA", "Responder de forma positiva y segura", "Prompt", "Sí", "Confirmado", "Regla de respuesta."),
    ("Lenguaje", "1.23", "Muletillas prohibidas", "systemPrompt", "texto: LENGUAJE.muletillas", "Evitar en resumen, en conclusión, aprovechar al máximo y exageraciones", "Prompt", "Sí", "Confirmado", "Regla de estilo."),
    ("Conversación", "1.24", "Prevención de eco", "systemPrompt", "texto: CONVERSACION.eco", "Evitar repetir las respuestas del usuario", "Prompt", "Sí", "Confirmado", "Regla conversacional."),
    ("Conversación", "1.25", "Confirmaciones breves habilitadas", "backchannel", "boolean", "", "Campo API", "Sí", "Pendiente", "Definir Sí/No."),
    ("Conversación", "1.26", "Palabras de confirmación", "backchannelWords", "array<string>", "", "Campo API", "Sí", "Pendiente", "Aplican únicamente si backchannel=true."),
    ("STT", "1.27", "Modelo de transcripción", "transcriptionModel", "string", "", "Campo API", "Sí", "Pendiente", "Separado del idioma."),
    ("Inactividad", "1.28", "Tiempo antes del mensaje", "idleTimeout", "number (segundos)", "", "Campo API", "Sí", "Pendiente", "Definir número de segundos."),
    ("Inactividad", "1.29", "Mensajes de inactividad", "idleMessages", "array<string>", "¿Hola, sigues ahí? Aún estoy en línea; por favor dime si sigues interesado", "Campo API", "Sí", "Confirmado", "Separar cada mensaje como elemento del arreglo."),
    ("Inactividad", "1.30", "Máximo de mensajes", "maxIdleMessages", "integer", 2, "Campo API", "Sí", "Confirmado", "Entero mayor o igual a cero."),
    ("Inactividad", "1.31", "Tiempo máximo de silencio", "silenceTimeout", "number (segundos)", "", "Campo API", "Sí", "Pendiente", "No estaba separado en el PDF."),
    ("Contexto", "2.1", "Origen del contacto", "systemPrompt", "texto: CONTEXTO.origen", "Lead proveniente de una lista cargada en CRM", "Prompt + integración", "Parcial", "Confirmado", "La carga real del contacto depende del CRM o campaña."),
    ("Contexto", "2.2", "Tipo de interacción", "systemPrompt", "texto: CONTEXTO.tipo", "Entrante y saliente", "Prompt", "Sí", "Pendiente", "Confirmar porque la celda Respuesta estaba vacía."),
    ("Contexto", "2.3", "Motivo del contacto", "systemPrompt", "texto: CONTEXTO.motivo", "Seguimiento a un lead que no agendó", "Prompt", "Sí", "Pendiente", "Tomado del ejemplo; confirmar como respuesta definitiva."),
    ("Contexto", "2.4", "Horario permitido", "smart_flow_id / campaña", "regla externa", "Lunes a viernes, 08:00–20:00", "Automatización", "No", "Pendiente", "La ejecución horaria no la controla systemPrompt."),
    ("Contexto", "2.5", "Zona horaria", "systemPrompt / flujo", "IANA timezone", "America/Bogota", "Prompt + integración", "Parcial", "Pendiente", "La programación efectiva pertenece al flujo o campaña."),
    ("Contexto", "2.6", "Política de tratamiento de datos", "systemPrompt", "texto: DATOS.politica", "No recopilar información; brindar soporte y solución", "Prompt", "Sí", "Pendiente", "Tomado del ejemplo; confirmar. Corrige el ID 2.5 repetido."),
    ("Objetivos", "3.1", "Objetivo principal 1", "systemPrompt", "texto: OBJETIVOS.principal_1", "Aclarar satisfactoriamente las dudas iniciales", "Prompt", "Sí", "Confirmado", "No existe como columna independiente."),
    ("Objetivos", "3.2", "Objetivo principal 2", "systemPrompt", "texto: OBJETIVOS.principal_2", "Determinar si procede contacto humano o si el usuario desiste", "Prompt + herramienta", "Parcial", "Confirmado", "La transferencia efectiva requiere herramienta o flujo."),
    ("Objetivos", "3.3", "Objetivo secundario", "analysisPrompt", "string", "Identificar motivos de la decisión del usuario", "Campo API", "Sí", "Confirmado", "Corrige el salto de 3.2 a 3.4."),
    ("Evaluación", "3.4", "Instrucciones de evaluación", "successEvaluationPrompt", "string", "Puntuar, confirmar objetivos cumplidos, temas relevantes y emoción", "Campo API", "Sí", "Confirmado", "Instrucción de evaluación posterior."),
    ("Evaluación", "3.5", "Rúbrica de evaluación", "successEvaluationRubric", "string", "Escala de 1 a 10", "Campo API", "Sí", "Confirmado", "Separada del prompt de evaluación."),
    ("Buzón", "4.1", "Buzón de voz habilitado", "voicemail", "boolean", True, "Campo API", "Sí", "Confirmado", "Resultado del PDF: Sí."),
    ("Buzón", "4.2", "Modo de buzón", "voicemailMode", "string", "hangup", "Campo API", "Sí", "Pendiente", "Confirmar comportamiento deseado después del mensaje."),
    ("Buzón", "4.3", "Mensaje de buzón", "voicemailMessage", "string", "Hola, te llamamos de Innovate... Gracias y feliz tarde", "Campo API", "Sí", "Confirmado", "Reemplazar los puntos suspensivos por el texto final aprobado."),
    ("Transferencia", "4.4", "Transferencia habilitada", "systemPrompt + tools", "regla + herramienta", False, "Prompt + integración", "Parcial", "Confirmado", "El PDF indica No."),
    ("Transferencia", "4.5", "Flujo o destino de transferencia", "smart_flow_id", "uuid (solo lectura)", "No aplica", "Automatización", "No", "No aplica", "No hay parámetro simple de pool en assistants."),
    ("Datos", "5.1", "Preguntas de la conversación", "systemPrompt", "texto: CUESTIONARIO", "Confirmar nombre y realizar tres preguntas operativas", "Prompt", "Sí", "Confirmado", "Corrige el ID 4.1 repetido del PDF."),
    ("Datos", "5.2", "Instrucciones de extracción", "structuredDataPrompt", "string", "Recopilar nombre, calificación, estado y respuestas", "Campo API", "Sí", "Confirmado", "Describe la extracción al finalizar."),
    ("Datos", "5.3", "Esquema de variables", "dataSchema", "array<object>", "nombre_usuario; calificacion_llamada; estado_interaccion; respuesta_pregunta_1..3", "Campo API", "Sí", "Confirmado", "Cada propiedad requiere propName, propType y description."),
    ("Webhook", "5.4", "URL de procesamiento posterior", "serverURL", "url", "", "Campo API + servicio", "Parcial", "Pendiente", "No incluir secretos; requiere endpoint operativo."),
    ("Webhook", "5.5", "Eventos enviados", "serverMessages", "array<string>", "structured-data; success-evaluation; summary", "Campo API", "Sí", "Pendiente", "Validar nombres exactos admitidos por la plataforma."),
]


CATALOG = [
    ("name", "string", "Sí", "Identificación", "Nombre del agente"),
    ("initialMessage", "string", "Sí", "Conversación", "Mensaje inicial"),
    ("systemPrompt", "string", "Sí", "Conversación", "Instrucciones completas"),
    ("multi_prompt", "array<object>", "Alto riesgo", "Conversación", "Mantener sincronizado con systemPrompt"),
    ("language", "locale", "Sí", "STT", "Idioma principal"),
    ("voice", "integer (FK)", "Sí", "TTS", "ID de voz"),
    ("ttsProvider", "integer (FK)", "Sí", "TTS", "Proveedor de síntesis"),
    ("ttsModel", "integer|string|null", "Sí", "TTS", "Modelo de síntesis"),
    ("voiceCustomSettings", "object", "Sí", "TTS", "Velocidad, estabilidad y claridad"),
    ("transcriptionModel", "string", "Sí", "STT", "Modelo de transcripción"),
    ("llmProvider", "integer (FK)", "Sí", "LLM", "Proveedor del modelo"),
    ("llmModel", "integer (FK)", "Sí", "LLM", "Modelo"),
    ("llmTemperature", "number", "Sí", "LLM", "Temperatura; validar rango en TracerCall"),
    ("maxTokens", "integer", "Sí", "LLM", "Máximo de tokens"),
    ("idleMessages", "array<string>", "Sí", "Inactividad", "Mensajes de inactividad"),
    ("idleTimeout", "number", "Sí", "Inactividad", "Tiempo antes del mensaje"),
    ("maxIdleMessages", "integer", "Sí", "Inactividad", "Número máximo"),
    ("silenceTimeout", "number", "Sí", "Inactividad", "Tiempo máximo de silencio"),
    ("backchannel", "boolean", "Sí", "Conversación", "Activa confirmaciones breves"),
    ("backchannelWords", "array<string>", "Sí", "Conversación", "Frases de confirmación"),
    ("voicemail", "boolean", "Sí", "Buzón", "Buzón habilitado"),
    ("voicemailMode", "string", "Sí", "Buzón", "Comportamiento del buzón"),
    ("voicemailMessage", "string", "Sí", "Buzón", "Mensaje del buzón"),
    ("analysisPrompt", "string", "Sí", "Postproceso", "Análisis y resumen"),
    ("successEvaluationPrompt", "string", "Sí", "Postproceso", "Evaluación"),
    ("successEvaluationRubric", "string", "Sí", "Postproceso", "Rúbrica"),
    ("structuredDataPrompt", "string", "Sí", "Postproceso", "Instrucciones de extracción"),
    ("dataSchema", "array<object>", "Sí", "Postproceso", "Variables estructuradas"),
    ("serverURL", "url", "Sí", "Webhook", "Destino del webhook"),
    ("serverMessages", "array<string>", "Sí", "Webhook", "Eventos enviados"),
    ("enabled_channels", "array<string>", "Sí", "Canal", "Canales habilitados"),
    ("smart_flow_id", "uuid", "No: solo lectura", "Automatización", "Flujo asociado"),
    ("knowledgeBase", "array<string>", "No: solo lectura", "Conocimiento", "Archivos asociados"),
    ("mcp_servers", "array<object>", "Restringir", "Herramientas", "No exponer ni editar tokens desde la ficha"),
    ("additional_settings", "object", "Validar", "Avanzado", "VAD, voz, fallback e interrupciones"),
]


EXAMPLES = {
    "name": "Agente Demo Innovate",
    "enabled_channels": '["voice"]',
    "initialMessage": "Hola, soy Andrés, agente especializado de Innovate. ¿En qué puedo orientarle?",
    "language": "es-419",
    "voice": "472",
    "ttsProvider": "2",
    "ttsModel": "null",
    "voiceCustomSettings.speed": "1.0",
    "llmTemperature": "0.3",
    "backchannel": "true",
    "backchannelWords": '["Claro", "Entiendo", "Sí"]',
    "transcriptionModel": "nova-3",
    "idleTimeout": "7",
    "idleMessages": '["¿Hola, sigue ahí?", "Aún estoy en línea"]',
    "maxIdleMessages": "2",
    "silenceTimeout": "30",
    "analysisPrompt": "Resume la interacción e identifica el motivo de la decisión del usuario.",
    "successEvaluationPrompt": "Evalúa de 1 a 10 si se cumplieron los objetivos de la interacción.",
    "successEvaluationRubric": "Escala de 1 a 10",
    "voicemail": "true",
    "voicemailMode": "hangup",
    "voicemailMessage": "Hola, le llamamos de Innovate. Nos comunicaremos nuevamente. Gracias.",
    "structuredDataPrompt": "Extrae el nombre, la calificación y las respuestas del cuestionario.",
    "dataSchema": '[{"propName":"nombre_usuario","propType":"string","description":"Nombre confirmado"}]',
    "serverURL": "https://backend.ejemplo.com/webhooks/sophia",
    "serverMessages": '["structured-data", "success-evaluation", "summary"]',
    "smart_flow_id / campaña": "Configurar la franja 08:00-20:00 en el flujo de llamadas",
    "systemPrompt / flujo": "Todas las horas se interpretan en America/Bogota.",
    "systemPrompt + tools": "Si se autoriza transferencia, ejecuta la herramienta transferir_a_asesor.",
    "smart_flow_id": "UUID del flujo configurado en SophIA",
}


def example_for(field):
    parameter = field[3]
    if parameter in EXAMPLES:
        return EXAMPLES[parameter]
    if parameter == "systemPrompt":
        label = field[2].lower()
        value = field[5]
        if value not in (None, ""):
            return f"Instrucción: {label}: {value}."
        return f"Instrucción: defina {label}."
    return "Completar según catálogo de la plataforma"


def style_title(ws, title: str, end_col: int):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_col)
    cell = ws.cell(1, 1, title)
    cell.fill = PatternFill("solid", fgColor="0B7A75")
    cell.font = Font(name="Aptos Display", size=16, bold=True, color="FFFFFF")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30


def style_header(ws, row: int, columns: int):
    for cell in ws[row][:columns]:
        cell.fill = PatternFill("solid", fgColor="17324D")
        cell.font = Font(name="Aptos", size=10, bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 34


def setup_print(ws, print_area: str):
    ws.print_area = print_area
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=0.2, right=0.2, top=0.35, bottom=0.35, header=0.15, footer=0.15)


OUT_DIR.mkdir(parents=True, exist_ok=True)
shutil.copy2(SOURCE, OUTPUT)
wb = load_workbook(OUTPUT)
if "Agente" in wb.sheetnames:
    wb["Agente"].title = "Ficha original"

for name in ["Ficha homologada", "Parámetros SophIA", "Guía de publicación"]:
    if name in wb.sheetnames:
        del wb[name]

ws = wb.create_sheet("Ficha homologada", 0)
style_title(ws, "FICHA MAESTRA HOMOLOGADA · AGENTE CONVERSACIONAL", 11)
ws.merge_cells("A2:K2")
ws["A2"] = "Complete la columna Valor. Los nombres técnicos coinciden con los parámetros observados en SophIA API v2."
ws["A2"].fill = PatternFill("solid", fgColor="D9F0ED")
ws["A2"].font = Font(name="Aptos", italic=True, color="17324D")
ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
headers = ["Sección", "ID", "Nombre homologado", "Parámetro SophIA", "Tipo de dato", "Valor", "Ejemplo", "Mecanismo", "Editable por API", "Estado", "Validación / observaciones"]
for col, value in enumerate(headers, 1):
    ws.cell(4, col, value)
style_header(ws, 4, len(headers))
for r, values in enumerate(FIELDS, 5):
    expanded = values[:6] + (example_for(values),) + values[6:]
    for c, value in enumerate(expanded, 1):
        ws.cell(r, c, value)
        ws.cell(r, c).alignment = Alignment(vertical="top", wrap_text=True)
        ws.cell(r, c).font = Font(name="Aptos", size=10)
    ws.cell(r, 6).fill = PatternFill("solid", fgColor="FFF9E8")
    ws.cell(r, 7).fill = PatternFill("solid", fgColor="EEF6FF")
    ws.row_dimensions[r].height = 44

last = 4 + len(FIELDS)
thin = Side(style="thin", color="D9E2EA")
for row in ws.iter_rows(min_row=4, max_row=last, min_col=1, max_col=11):
    for cell in row:
        cell.border = Border(bottom=thin)

widths = {"A": 16, "B": 9, "C": 31, "D": 32, "E": 25, "F": 44, "G": 48, "H": 22, "I": 16, "J": 16, "K": 58}
for col, width in widths.items():
    ws.column_dimensions[col].width = width
ws.freeze_panes = "A5"
ws.auto_filter.ref = f"A4:J{last}"
ws.sheet_view.showGridLines = False
ws.sheet_properties.pageSetUpPr.fitToPage = True
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0
setup_print(ws, f"A1:K{last}")
ws.sheet_properties.outlinePr.summaryBelow = True

status_dv = DataValidation(type="list", formula1='"Confirmado,Pendiente,No aplica"', allow_blank=False)
ws.add_data_validation(status_dv)
status_dv.add(f"J5:J{last}")

green = PatternFill("solid", fgColor="DCFCE7")
yellow = PatternFill("solid", fgColor="FEF3C7")
gray = PatternFill("solid", fgColor="E5E7EB")
ws.conditional_formatting.add(f"J5:J{last}", FormulaRule(formula=["$J5=\"Confirmado\""], fill=green))
ws.conditional_formatting.add(f"J5:J{last}", FormulaRule(formula=["$J5=\"Pendiente\""], fill=yellow))
ws.conditional_formatting.add(f"J5:J{last}", FormulaRule(formula=["$J5=\"No aplica\""], fill=gray))

table = Table(displayName="FichaHomologada", ref=f"A4:K{last}")
table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showFirstColumn=False, showLastColumn=False)
ws.add_table(table)

pc = wb.create_sheet("Parámetros SophIA", 1)
style_title(pc, "CATÁLOGO DE PARÁMETROS SOPHIA", 5)
pc_headers = ["Parámetro exacto", "Tipo", "Escritura API", "Grupo", "Uso"]
for col, value in enumerate(pc_headers, 1):
    pc.cell(3, col, value)
style_header(pc, 3, 5)
for r, values in enumerate(CATALOG, 4):
    for c, value in enumerate(values, 1):
        pc.cell(r, c, value)
        pc.cell(r, c).alignment = Alignment(vertical="top", wrap_text=True)
    pc.row_dimensions[r].height = 30
pc_last = 3 + len(CATALOG)
for col, width in {"A": 34, "B": 24, "C": 20, "D": 20, "E": 58}.items():
    pc.column_dimensions[col].width = width
pc.freeze_panes = "A4"
pc.auto_filter.ref = f"A3:E{pc_last}"
pc.sheet_view.showGridLines = False
setup_print(pc, f"A1:E{pc_last}")
pc_table = Table(displayName="CatalogoParametros", ref=f"A3:E{pc_last}")
pc_table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
pc.add_table(pc_table)

guide = wb.create_sheet("Guía de publicación", 2)
style_title(guide, "REGLAS DE IMPORTACIÓN Y PUBLICACIÓN", 5)
guide_headers = ["Paso", "Regla", "Responsable", "Resultado", "Obligatorio"]
for col, value in enumerate(guide_headers, 1):
    guide.cell(3, col, value)
style_header(guide, 3, 5)
steps = [
    (1, "Extraer exclusivamente los valores de la ficha; ejemplos no confirmados quedan Pendientes.", "TracerCall", "Borrador validado", "Sí"),
    (2, "Detener la publicación si hay valores vacíos, duplicados o contradictorios.", "Validador", "Lista de pendientes", "Sí"),
    (3, "Resolver nombres comerciales de voz, TTS y LLM a IDs internos vigentes.", "API", "Referencias válidas", "Sí"),
    (4, "Construir systemPrompt por bloques sin inventar parámetros de plataforma.", "Generador", "Prompt versionado", "Sí"),
    (5, "Respaldar GET /assistants?id=<UUID> antes de cualquier escritura.", "Backend", "Snapshot recuperable", "Sí"),
    (6, "Actualizar únicamente por id exacto y mediante una lista permitida de campos.", "Backend", "PUT controlado", "Sí"),
    (7, "Releer el agente, comparar valores y hash, y registrar auditoría.", "Backend", "Publicación verificada", "Sí"),
    (8, "Mantener systemPrompt y multi_prompt sincronizados o bloquear edición desde consola.", "Backend / operación", "Sin sobrescritura accidental", "Sí"),
]
for r, values in enumerate(steps, 4):
    for c, value in enumerate(values, 1):
        guide.cell(r, c, value)
        guide.cell(r, c).alignment = Alignment(vertical="top", wrap_text=True)
    guide.row_dimensions[r].height = 40

guide["A14"] = "Resumen de estado"
guide["A14"].font = Font(bold=True, color="FFFFFF")
guide["A14"].fill = PatternFill("solid", fgColor="17324D")
guide.merge_cells("A14:B14")
guide["A15"], guide["B15"] = "Confirmados", f'=COUNTIF(\'Ficha homologada\'!$J$5:$J${last},"Confirmado")'
guide["A16"], guide["B16"] = "Pendientes", f'=COUNTIF(\'Ficha homologada\'!$J$5:$J${last},"Pendiente")'
guide["A17"], guide["B17"] = "No aplica", f'=COUNTIF(\'Ficha homologada\'!$J$5:$J${last},"No aplica")'
for col, width in {"A": 12, "B": 72, "C": 24, "D": 34, "E": 15}.items():
    guide.column_dimensions[col].width = width
guide.freeze_panes = "A4"
guide.sheet_view.showGridLines = False
setup_print(guide, "A1:E17")

original = wb["Ficha original"]
original.sheet_view.showGridLines = False
original.sheet_state = "visible"
if "MCP" in wb.sheetnames:
    wb["MCP"].sheet_state = "visible"

wb.active = 0
wb.calculation.fullCalcOnLoad = True
wb.calculation.forceFullCalc = True
wb.save(OUTPUT)
print(OUTPUT)
